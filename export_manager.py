from __future__ import annotations

from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
import json
import re

import numpy as np
import pandas as pd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

from pl_core import CompletedFit, FitSettings, build_fit_curve_data, result_to_dict


WINDOWS_FORBIDDEN = re.compile(r'[<>:"/\\|?*\x00-\x1f]')
EXCEL_FORBIDDEN = re.compile(r'[\[\]:*?/\\\x00-\x1f]')
WINDOWS_RESERVED = {
    "CON", "PRN", "AUX", "NUL",
    *(f"COM{index}" for index in range(1, 10)),
    *(f"LPT{index}" for index in range(1, 10)),
}


@dataclass(frozen=True)
class ExportOptions:
    summary_excel: bool = True
    individual_csv: bool = True
    individual_excel: bool = True
    figures_png: bool = True
    figures_pdf: bool = True
    legacy_curve_sheets: bool = False
    timestamped_folder: bool = True


@dataclass
class _ExportEntry:
    record_id: str
    output_name: str
    source_filename: str
    source_full_path: str
    condition: str
    sample_id: str
    temperature: Any
    completed_fit: CompletedFit | None
    fit_status: str
    failure_reason: str = ""
    csv_filename: str = ""
    sheet_name: str = ""

    @property
    def spectrum(self):
        return self.completed_fit.spectrum if self.completed_fit is not None else None

    @property
    def result(self):
        return self.completed_fit.result if self.completed_fit is not None else None


def _coerce_options(options: ExportOptions | dict[str, bool] | None) -> ExportOptions:
    if options is None:
        return ExportOptions()
    if isinstance(options, ExportOptions):
        return options
    return ExportOptions(**options)


def sanitize_windows_filename(value: str, *, maximum_length: int = 150) -> str:
    """Return a valid basename while preserving the user's Output Name elsewhere."""
    name = WINDOWS_FORBIDDEN.sub("_", str(value)).strip().rstrip(". ")
    if name.lower().endswith(".csv"):
        name = name[:-4].rstrip(". ")
    name = name[:maximum_length].rstrip(". ") or "spectrum"
    device_stem, separator, remainder = name.partition(".")
    if device_stem.upper() in WINDOWS_RESERVED:
        name = f"{device_stem}_{separator}{remainder}" if separator else f"{device_stem}_"
    return name


def unique_csv_filename(output_name: str, used_names: set[str]) -> str:
    base = sanitize_windows_filename(output_name)
    candidate = f"{base}.csv"
    suffix = 2
    while candidate.casefold() in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{base[:150-len(suffix_text)].rstrip()}{suffix_text}.csv"
        suffix += 1
    used_names.add(candidate.casefold())
    return candidate


def unique_excel_sheet_name(value: str, used_names: set[str]) -> str:
    base = EXCEL_FORBIDDEN.sub("_", str(value)).strip().strip("'") or "Fit"
    base = base[:31]
    candidate = base
    suffix = 2
    while candidate.casefold() in used_names:
        suffix_text = f"_{suffix}"
        candidate = f"{base[:31-len(suffix_text)].rstrip()}{suffix_text}"
        suffix += 1
    used_names.add(candidate.casefold())
    return candidate


def fit_curve_dataframe(result) -> pd.DataFrame:
    """Build an Origin-friendly table exclusively from shared saved curve arrays."""
    curves = build_fit_curve_data(result)
    data = {
        "Photon_Energy_eV": curves.photon_energy_ev,
        "Wavelength_nm": curves.wavelength_nm,
        "Fit_Input_Intensity": curves.raw_intensity,
    }
    if curves.baseline is not None:
        data["Baseline"] = curves.baseline
    for index, component in enumerate(curves.components, 1):
        data[f"Peak_{index}"] = component
        if curves.baseline is not None:
            data[f"Peak_{index}_With_Baseline"] = component + curves.baseline
    data["Total_Fit"] = curves.total_fit
    data["Residual"] = curves.residual
    return pd.DataFrame(data)


def _coerce_completed_fits(rows: Iterable, *, records_supplied: bool) -> list[CompletedFit]:
    completed: list[CompletedFit] = []
    for index, row in enumerate(rows):
        if isinstance(row, CompletedFit):
            item = row
        elif isinstance(row, (tuple, list)) and len(row) == 4:
            record_id, output_name, spectrum, result = row
            item = CompletedFit(str(record_id), str(output_name), spectrum, result)
        elif isinstance(row, (tuple, list)) and len(row) == 3 and not records_supplied:
            output_name, spectrum, result = row
            item = CompletedFit(f"legacy-export-{index + 1}", str(output_name), spectrum, result)
        else:
            raise ValueError(
                "Exports with GUI records require CompletedFit values carrying record_id; "
                "filename and Output Name matching are intentionally unsupported."
            )
        completed.append(item)
    ids = [item.record_id for item in completed]
    if len(ids) != len(set(ids)):
        raise ValueError("Completed fit record_id values must be unique.")
    return completed


def _records_for_legacy_rows(completed: list[CompletedFit]) -> list[dict[str, Any]]:
    return [
        {
            "record_id": item.record_id,
            "export": True,
            "code": item.output_name,
            "path": item.spectrum.source_name,
            "condition": item.spectrum.condition,
            "sample_id": item.spectrum.sample_id,
            "temperature": item.spectrum.temperature_k,
            "status": "OK",
        }
        for item in completed
    ]


def _build_export_selection(rows, records=None) -> tuple[list[_ExportEntry], int, int]:
    records_supplied = records is not None
    completed = _coerce_completed_fits(rows, records_supplied=records_supplied)
    record_list = [dict(record) for record in records] if records_supplied else _records_for_legacy_rows(completed)
    record_ids = [str(record.get("record_id", "")).strip() for record in record_list]
    if any(not record_id for record_id in record_ids):
        raise ValueError("Every export record must have a non-empty record_id.")
    if len(record_ids) != len(set(record_ids)):
        raise ValueError("Export record_id values must be unique.")

    completed_by_id = {item.record_id: item for item in completed}
    entries: list[_ExportEntry] = []
    skipped = 0
    for record in record_list:
        if not record.get("export", True):
            skipped += 1
            continue
        record_id = str(record["record_id"])
        item = completed_by_id.get(record_id)
        if item is not None:
            spectrum = item.spectrum
            entries.append(
                _ExportEntry(
                    record_id=record_id,
                    output_name=str(record.get("code", item.output_name)),
                    source_filename=spectrum.source_name,
                    source_full_path=str(record.get("path", "")),
                    condition=spectrum.condition,
                    sample_id=spectrum.sample_id,
                    temperature=spectrum.temperature_k,
                    completed_fit=item,
                    fit_status="Successful",
                )
            )
            continue
        status = str(record.get("status", ""))
        entries.append(
            _ExportEntry(
                record_id=record_id,
                output_name=str(record.get("code", "")),
                source_filename=Path(str(record.get("path", ""))).name,
                source_full_path=str(record.get("path", "")),
                condition=str(record.get("condition", "")),
                sample_id=str(record.get("sample_id", "")),
                temperature=record.get("temperature", ""),
                completed_fit=None,
                fit_status="Failed" if status.casefold() == "failed" else (status or "Not fitted"),
                failure_reason=str(record.get("failure_reason", "")),
            )
        )
    return entries, skipped, len(completed)


def _settings_for_result(result) -> FitSettings | None:
    settings = getattr(result, "fit_settings", None)
    return settings if isinstance(settings, FitSettings) else None


def _fit_metadata_dict(entry: _ExportEntry) -> dict[str, Any]:
    metadata: dict[str, Any] = {
        "Record_ID": entry.record_id,
        "Output_Name": entry.output_name,
        "Source_Filename": entry.source_filename,
        "Source_Full_Path": entry.source_full_path,
        "Condition": entry.condition,
        "Sample_ID": entry.sample_id,
        "Temperature_K": entry.temperature,
    }
    result = entry.result
    if result is None:
        return metadata
    settings = _settings_for_result(result)
    metadata.update(
        {
            "Model": result.model,
            "Number_of_Peaks": result.n_peaks,
            "Requested_Model": settings.requested_model if settings else "Unknown (legacy result)",
            "Requested_Number_of_Peaks": settings.requested_n_peaks if settings else "Unknown (legacy result)",
            "Baseline_Mode": settings.baseline_mode if settings else (result.baseline_mode or "Unknown (legacy result)"),
            "Jacobian_Corrected": settings.jacobian_corrected if settings else None,
            "Intensity_Domain": settings.intensity_domain if settings else "Unknown (legacy result)",
            "Fit_Input_Intensity_Definition": "Fitting-domain intensity; may be Jacobian-corrected when enabled",
            "Fit_Range_eV": list(settings.fit_range_ev) if settings else [float(np.min(result.x_ev)), float(np.max(result.x_ev))],
            "Instrument_FWHM_meV": settings.instrument_fwhm_mev if settings else None,
            "Automatic_Model_Selection": settings.automatic_model_selection if settings else None,
            "R2": result.r_squared,
            "Adjusted_R2": result.adjusted_r_squared,
            "RMSE": result.rmse,
            "AICc": result.aicc,
            "BIC": result.bic,
        }
    )
    for peak in result.peaks:
        prefix = f"Peak_{peak.peak_index}"
        metadata.update(
            {
                f"{prefix}_Energy_eV": peak.center_ev,
                f"{prefix}_Wavelength_nm": peak.center_nm,
                f"{prefix}_FWHM_meV": peak.fwhm_mev,
                f"{prefix}_Height": peak.height,
                f"{prefix}_Area": peak.area,
                f"{prefix}_Area_Fraction": peak.area_fraction,
            }
        )
    return metadata


def _excel_value(value: Any) -> Any:
    if isinstance(value, (list, tuple, dict)):
        return json.dumps(value, ensure_ascii=False)
    return value


def _index_rows(entries: list[_ExportEntry]) -> list[dict[str, Any]]:
    output = []
    for entry in entries:
        row = _fit_metadata_dict(entry)
        row.update(
            {
                "Export_Sheet_Name": entry.sheet_name,
                "CSV_Filename": entry.csv_filename,
                "Fit_Status": entry.fit_status,
                "Failure_Reason": entry.failure_reason,
            }
        )
        output.append(row)
    return output


def _style_header(row) -> None:
    fill = PatternFill("solid", fgColor="1F4E78")
    for cell in row:
        cell.font = Font(color="FFFFFF", bold=True)
        cell.fill = fill


def _write_individual_excel(path: Path, entries: list[_ExportEntry]) -> None:
    index_rows = _index_rows(entries)
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(index_rows).map(_excel_value).to_excel(writer, sheet_name="Index", index=False)
        index_sheet = writer.sheets["Index"]
        _style_header(index_sheet[1])
        index_sheet.freeze_panes = "A2"
        index_sheet.auto_filter.ref = index_sheet.dimensions
        index_sheet.sheet_view.showGridLines = False
        for column_index in range(1, index_sheet.max_column + 1):
            index_sheet.column_dimensions[get_column_letter(column_index)].width = 24

        for entry in entries:
            if entry.completed_fit is None:
                continue
            metadata = [(name, _excel_value(value)) for name, value in _fit_metadata_dict(entry).items()]
            metadata_frame = pd.DataFrame(metadata, columns=["Metadata_Field", "Value"])
            metadata_frame.to_excel(writer, sheet_name=entry.sheet_name, index=False)
            curve_frame = fit_curve_dataframe(entry.result)
            curve_startrow = len(metadata_frame) + 2
            curve_frame.to_excel(writer, sheet_name=entry.sheet_name, startrow=curve_startrow, index=False)
            worksheet = writer.sheets[entry.sheet_name]
            _style_header(worksheet[1])
            _style_header(worksheet[curve_startrow + 1])
            worksheet.freeze_panes = f"A{curve_startrow + 2}"
            last_column = worksheet.cell(curve_startrow + 1, curve_frame.shape[1]).column_letter
            worksheet.auto_filter.ref = f"A{curve_startrow + 1}:{last_column}{curve_startrow + 1 + len(curve_frame)}"
            worksheet.sheet_view.showGridLines = False
            worksheet.column_dimensions["A"].width = 34
            worksheet.column_dimensions["B"].width = 36
            for column_index in range(3, curve_frame.shape[1] + 1):
                worksheet.column_dimensions[get_column_letter(column_index)].width = 24


def _collect_summary_rows(entries: list[_ExportEntry], tempfits) -> tuple[list[dict], list[dict]]:
    peak_rows = []
    for entry in entries:
        if entry.completed_fit is None:
            continue
        spectrum, result = entry.spectrum, entry.result
        for peak in result.peaks:
            peak_rows.append(
                {
                    "Record_ID": entry.record_id, "Label": entry.output_name,
                    "Condition": spectrum.condition, "Sample ID": spectrum.sample_id,
                    "Temperature_K": spectrum.temperature_k, "Model": result.model,
                    "N peaks": result.n_peaks, "Peak": peak.peak_index,
                    "Center_eV": peak.center_ev, "Center_nm": peak.center_nm,
                    "FWHM_meV": peak.fwhm_mev, "Height": peak.height,
                    "Area": peak.area, "Area_fraction": peak.area_fraction,
                    "R2": result.r_squared, "Adjusted_R2": result.adjusted_r_squared,
                    "RMSE": result.rmse, "AICc": result.aicc, "BIC": result.bic,
                    "Warnings": "; ".join(result.warnings),
                }
            )

    temp_rows = []
    for label, group in tempfits.items():
        for kind, result in group.items():
            for name, value in result.parameters.items():
                low, high = result.ci95[name]
                temp_rows.append(
                    {
                        "Label": label, "Analysis": kind, "Model": result.model,
                        "Parameter": name, "Value": value, "SE": result.errors[name],
                        "CI95_low": low, "CI95_high": high, "R2": result.r_squared,
                        "Adjusted_R2": result.adjusted_r_squared, "RMSE": result.rmse,
                        "AICc": result.aicc, "BIC": result.bic,
                        "Warnings": "; ".join(result.warnings),
                    }
                )
    return peak_rows, temp_rows


def _spectrum_metadata_rows(entries: list[_ExportEntry]) -> list[dict[str, Any]]:
    rows = []
    for entry in entries:
        if entry.completed_fit is None:
            continue
        source_metadata = entry.spectrum.metadata or {"": ""}
        for name, value in source_metadata.items():
            rows.append(
                {
                    "Record_ID": entry.record_id, "Output_Name": entry.output_name,
                    "Source": entry.source_filename, "Source_Full_Path": entry.source_full_path,
                    "Key": name, "Value": value,
                }
            )
    return rows


def _write_summary_excel(path: Path, entries: list[_ExportEntry], peak_rows: list[dict], temp_rows: list[dict], *, legacy_curve_sheets: bool) -> None:
    metadata_rows = _spectrum_metadata_rows(entries)
    used_sheet_names = {"peak_results", "temperature_fits", "spectrum_metadata"}
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(peak_rows).to_excel(writer, sheet_name="Peak_Results", index=False)
        pd.DataFrame(temp_rows).to_excel(writer, sheet_name="Temperature_Fits", index=False)
        pd.DataFrame(metadata_rows, columns=["Record_ID", "Output_Name", "Source", "Source_Full_Path", "Key", "Value"]).to_excel(writer, sheet_name="Spectrum_Metadata", index=False)
        if not legacy_curve_sheets:
            return
        for entry in entries:
            if entry.completed_fit is None:
                continue
            key = f"{entry.output_name}_{entry.temperature:g}K"
            fit_sheet = unique_excel_sheet_name(f"Fit_{key}", used_sheet_names)
            residual_sheet = unique_excel_sheet_name(f"Res_{key}", used_sheet_names)
            curves = build_fit_curve_data(entry.result)
            curve_data = {"Energy_eV": curves.photon_energy_ev, "Fit_Input_Intensity": curves.raw_intensity, "Total_fit": curves.total_fit}
            if curves.baseline is not None:
                curve_data["Baseline"] = curves.baseline
            for index, component in enumerate(curves.components, 1):
                curve_data[f"Peak_{index}"] = component
            pd.DataFrame(curve_data).to_excel(writer, sheet_name=fit_sheet, index=False)
            pd.DataFrame({"Energy_eV": curves.photon_energy_ev, "Residual": curves.residual}).to_excel(writer, sheet_name=residual_sheet, index=False)


def _timestamped_directory(base: Path) -> Path:
    stem = f"PL_Export_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
    candidate = base / stem
    suffix = 2
    while candidate.exists():
        candidate = base / f"{stem}_{suffix}"
        suffix += 1
    return candidate


def _safe_relative_target(root: Path, relative_name: str) -> Path | None:
    target = (root / relative_name).resolve()
    try:
        target.relative_to(root.resolve())
    except ValueError:
        return None
    return target


def _cleanup_previous_generated_files(root: Path) -> None:
    """Remove only files explicitly listed by this app's preceding manifest."""
    manifest_path = root / "Export_Manifest.json"
    if not manifest_path.is_file():
        return
    try:
        manifest = json.loads(manifest_path.read_text(encoding="utf-8"))
    except (OSError, ValueError, TypeError):
        return
    for relative_name in manifest.get("generated_files", []):
        target = _safe_relative_target(root, str(relative_name))
        if target is not None and target.is_file():
            target.unlink()


def _relative(path: Path, root: Path) -> str:
    return path.relative_to(root).as_posix()


def export_all(outdir, rows, tempfits, figures, *, records=None, options=None, app_version: str = "unknown"):
    """Export one internally consistent dataset selected solely by record_id."""
    options = _coerce_options(options)
    base = Path(outdir)
    base.mkdir(parents=True, exist_ok=True)
    out = _timestamped_directory(base) if options.timestamped_folder else base
    if not options.timestamped_folder:
        _cleanup_previous_generated_files(out)
    out.mkdir(parents=True, exist_ok=True)

    entries, skipped, successful_fit_count = _build_export_selection(rows, records)
    successful_entries = [entry for entry in entries if entry.completed_fit is not None]
    used_csv_names: set[str] = set()
    used_sheet_names = {"index"}
    for entry in successful_entries:
        if options.individual_csv:
            entry.csv_filename = unique_csv_filename(entry.output_name, used_csv_names)
        if options.individual_excel:
            entry.sheet_name = unique_excel_sheet_name(entry.output_name, used_sheet_names)

    generated: list[Path] = []
    peak_rows, temp_rows = _collect_summary_rows(entries, tempfits)
    if options.summary_excel:
        summary_path = out / "PL_batch_results.xlsx"
        _write_summary_excel(summary_path, entries, peak_rows, temp_rows, legacy_curve_sheets=options.legacy_curve_sheets)
        generated.append(summary_path)

    peak_csv = out / "peak_results.csv"
    temp_csv = out / "temperature_fit_results.csv"
    pd.DataFrame(peak_rows).to_csv(peak_csv, index=False, encoding="utf-8-sig")
    pd.DataFrame(temp_rows).to_csv(temp_csv, index=False, encoding="utf-8-sig")
    generated.extend((peak_csv, temp_csv))

    payload = {
        "spectra": [
            {
                "record_id": entry.record_id, "label": entry.output_name,
                "source_full_path": entry.source_full_path,
                "spectrum": {
                    "condition": entry.spectrum.condition, "sample_id": entry.spectrum.sample_id,
                    "temperature_k": entry.spectrum.temperature_k,
                    "source_name": entry.spectrum.source_name, "metadata": entry.spectrum.metadata,
                },
                "fit": result_to_dict(entry.result),
            }
            for entry in successful_entries
        ],
        "temperature_fits": {
            label: {name: result_to_dict(value) for name, value in group.items()}
            for label, group in tempfits.items()
        },
    }
    analysis_path = out / "analysis_results.json"
    analysis_path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    generated.append(analysis_path)

    csv_count = 0
    if options.individual_csv:
        csv_directory = out / "Individual_Fit_Data"
        csv_directory.mkdir(parents=True, exist_ok=True)
        for entry in successful_entries:
            csv_path = csv_directory / entry.csv_filename
            fit_curve_dataframe(entry.result).to_csv(csv_path, index=False, encoding="utf-8-sig", float_format="%.17g")
            generated.append(csv_path)
            csv_count += 1
        index_path = csv_directory / "Index.json"
        index_path.write_text(json.dumps(_index_rows(entries), ensure_ascii=False, indent=2), encoding="utf-8")
        generated.append(index_path)

    individual_excel_sheets = 0
    if options.individual_excel:
        individual_path = out / "Individual_Fitting_Data.xlsx"
        _write_individual_excel(individual_path, entries)
        generated.append(individual_path)
        individual_excel_sheets = len(successful_entries)

    if options.figures_png or options.figures_pdf:
        figure_directory = out / "Figures"
        figure_directory.mkdir(parents=True, exist_ok=True)
        for name, figure in figures.items():
            safe_name = sanitize_windows_filename(name)
            if options.figures_png:
                png_path = figure_directory / f"{safe_name}.png"
                figure.savefig(png_path, dpi=300)
                generated.append(png_path)
            if options.figures_pdf:
                pdf_path = figure_directory / f"{safe_name}.pdf"
                figure.savefig(pdf_path)
                generated.append(pdf_path)

    manifest_path = out / "Export_Manifest.json"
    failed_count = sum(entry.fit_status == "Failed" for entry in entries)
    not_fitted_count = sum(
        entry.completed_fit is None and entry.fit_status != "Failed" for entry in entries
    )
    manifest = {
        "export_timestamp": datetime.now().astimezone().isoformat(), "app_version": app_version,
        "record_count": len(records) if records is not None else len(entries),
        "successful_fit_count": successful_fit_count, "selected_export_count": len(entries),
        "successful_selected_fit_count": len(successful_entries), "failed_fit_count": failed_count,
        "not_fitted_count": not_fitted_count,
        "skipped_by_export_checkbox": skipped, "csv_count": csv_count,
        "source_files": [entry.source_full_path for entry in entries],
        "output_names": [entry.output_name for entry in entries],
        "generated_files": [_relative(path, out) for path in generated] + ["Export_Manifest.json"],
    }
    manifest_path.write_text(json.dumps(manifest, ensure_ascii=False, indent=2), encoding="utf-8")

    return {
        "output_directory": str(out), "successful_curve_exports": len(successful_entries),
        "csv_files": csv_count, "individual_excel_sheets": individual_excel_sheets,
        "index_rows": len(entries), "failed_fits": failed_count,
        "skipped_by_export_checkbox": skipped,
    }
