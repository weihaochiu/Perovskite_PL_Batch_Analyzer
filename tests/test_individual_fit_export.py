import json
from pathlib import Path
import tempfile
import unittest

import matplotlib
import numpy as np
import pandas as pd
from openpyxl import load_workbook

matplotlib.use("Agg")

from export_manager import ExportOptions, export_all, fit_curve_dataframe, sanitize_windows_filename
from pl_core import (
    CompletedFit,
    FitSettings,
    HC_EV_NM,
    PLSpectrum,
    PeakResult,
    SpectrumFitResult,
    TemperatureFitResult,
    build_fit_curve_data,
)
from plotting import selected_fit_figure


DIRECT_CSV = ExportOptions(
    summary_excel=False,
    individual_csv=True,
    individual_excel=False,
    figures_png=False,
    figures_pdf=False,
    timestamped_folder=False,
)
DIRECT_CURVES = ExportOptions(
    summary_excel=False,
    individual_csv=True,
    individual_excel=True,
    figures_png=False,
    figures_pdf=False,
    timestamped_folder=False,
)


def make_completed(
    record_id,
    output_name,
    *,
    source_path=None,
    n_peaks=1,
    model="Gaussian",
    baseline_mode="Constant",
    jacobian=False,
    marker=0.0,
):
    energy = np.linspace(1.35, 1.85, 16)
    baseline = np.linspace(1.0, 1.5, len(energy)) if baseline_mode != "None" else np.zeros_like(energy)
    components = []
    peaks = []
    for index in range(1, n_peaks + 1):
        center = 1.4 + index * 0.1
        component = (10.0 * index + marker) * np.exp(-4 * np.log(2) * ((energy - center) / 0.08) ** 2)
        components.append(component)
        peaks.append(
            PeakResult(
                peak_index=index,
                center_ev=center,
                center_nm=HC_EV_NM / center,
                fwhm_ev=0.08,
                fwhm_mev=80.0,
                height=10.0 * index + marker,
                area=float(np.trapezoid(component, energy)),
                area_fraction=1.0 / n_peaks,
                parameters={},
                errors={},
            )
        )
    total_fit = baseline + np.sum(components, axis=0)
    residual = np.linspace(-0.25, 0.25, len(energy))
    fit_input = total_fit + residual
    source_path = source_path or str(Path("input") / f"{output_name}.asc")
    source_name = Path(source_path).name
    spectrum = PLSpectrum(
        wavelength_nm=HC_EV_NM / energy,
        photon_energy_ev=energy,
        intensity_wavelength=fit_input.copy(),
        intensity_energy=fit_input.copy(),
        temperature_k=300.0,
        condition="C1",
        sample_id=f"sample-{record_id}",
        source_name=source_name,
        metadata={"Instrument": "Synthetic test fixture"},
        input_x_type="Wavelength (nm)",
    )
    settings = FitSettings(
        requested_model=model,
        requested_n_peaks=str(n_peaks),
        selected_model=model,
        selected_n_peaks=n_peaks,
        baseline_mode=baseline_mode,
        jacobian_corrected=jacobian,
        intensity_domain="Energy-domain Jacobian-corrected" if jacobian else "Wavelength-domain input",
        fit_range_ev=(float(energy.min()), float(energy.max())),
        instrument_fwhm_mev=5.0,
        automatic_model_selection=False,
    )
    result = SpectrumFitResult(
        temperature_k=300.0,
        model=model,
        n_peaks=n_peaks,
        peaks=peaks,
        x_ev=energy,
        y_raw=fit_input,
        y_fit=total_fit,
        components=components,
        baseline=baseline,
        residual=residual,
        r_squared=0.999,
        adjusted_r_squared=0.998,
        rmse=0.1,
        reduced_chi2=0.01,
        aic=1.0,
        aicc=2.0,
        bic=3.0,
        warnings=[],
        baseline_mode=baseline_mode,
        fit_settings=settings,
    )
    completed = CompletedFit(record_id, output_name, spectrum, result)
    record = {
        "record_id": record_id,
        "use": True,
        "export": True,
        "condition": spectrum.condition,
        "sample_id": spectrum.sample_id,
        "temperature": spectrum.temperature_k,
        "direction": "Isothermal",
        "code": output_name,
        "path": source_path,
        "x": "Wavelength_nm",
        "y": "PL_Intensity",
        "status": "OK",
    }
    return completed, record


def run_directory(report):
    return Path(report["output_directory"])


class TestIndividualFitExport(unittest.TestCase):
    def test_curve_columns_components_and_gui_arrays_share_saved_data(self):
        completed, _ = make_completed("id-1", "one", n_peaks=3)
        result = completed.result
        curves = build_fit_curve_data(result)
        frame = fit_curve_dataframe(result)
        expected = ["Photon_Energy_eV", "Wavelength_nm", "Fit_Input_Intensity", "Baseline"]
        for index in range(1, 4):
            expected.extend([f"Peak_{index}", f"Peak_{index}_With_Baseline"])
        expected.extend(["Total_Fit", "Residual"])
        self.assertEqual(frame.columns.tolist(), expected)
        np.testing.assert_allclose(frame["Fit_Input_Intensity"] - frame["Total_Fit"], frame["Residual"])
        np.testing.assert_allclose(curves.total_fit, curves.baseline + np.sum(curves.components, axis=0))
        for index, component in enumerate(curves.components, 1):
            np.testing.assert_allclose(frame[f"Peak_{index}_With_Baseline"], component + curves.baseline)

        figure = selected_fit_figure(result)
        plot_lines = {line.get_label(): line.get_ydata() for line in figure.axes[0].lines}
        np.testing.assert_allclose(plot_lines["Raw"], curves.raw_intensity)
        np.testing.assert_allclose(plot_lines["Total fit"], curves.total_fit)
        np.testing.assert_allclose(plot_lines["Baseline"], curves.baseline)
        for index, component in enumerate(curves.components, 1):
            np.testing.assert_allclose(plot_lines[f"Peak {index}"], component)
        figure.clear()

    def test_no_baseline_omits_baseline_and_display_columns(self):
        completed, _ = make_completed("id-none", "none", baseline_mode="None")
        columns = fit_curve_dataframe(completed.result).columns.tolist()
        self.assertNotIn("Baseline", columns)
        self.assertFalse(any(column.endswith("_With_Baseline") for column in columns))

    def test_all_models_and_peak_counts_export_pure_components(self):
        for model in ["Gaussian", "Lorentzian", "Pseudo-Voigt", "Voigt"]:
            for n_peaks in [1, 2, 3]:
                completed, _ = make_completed("id", "multi", n_peaks=n_peaks, model=model)
                frame = fit_curve_dataframe(completed.result)
                pure = [column for column in frame if column.startswith("Peak_") and not column.endswith("_With_Baseline")]
                self.assertEqual(pure, [f"Peak_{index}" for index in range(1, n_peaks + 1)])

    def test_79_records_and_79_fits_export_79_csv_and_sheets_with_identity(self):
        rows, records = [], []
        for index in range(79):
            completed, record = make_completed(f"record-{index:03d}", f"C{index + 1}", source_path=f"D:/Run/source_{index + 1}.asc", marker=index)
            rows.append(completed)
            records.append(record)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, rows, {}, {}, records=records, options=DIRECT_CURVES)
            out = run_directory(report)
            self.assertEqual(report["successful_curve_exports"], 79)
            self.assertEqual(len(list((out / "Individual_Fit_Data").glob("*.csv"))), 79)
            workbook = load_workbook(out / "Individual_Fitting_Data.xlsx", read_only=True)
            self.assertEqual(len(workbook.sheetnames), 80)
            workbook.close()
            index = json.loads((out / "Individual_Fit_Data" / "Index.json").read_text(encoding="utf-8"))
            self.assertEqual({row["Record_ID"] for row in index}, {record["record_id"] for record in records})
            self.assertEqual({row["Output_Name"] for row in index}, {record["code"] for record in records})
            manifest = json.loads((out / "Export_Manifest.json").read_text(encoding="utf-8"))
            self.assertEqual(manifest["record_count"], 79)
            self.assertEqual(manifest["successful_fit_count"], 79)
            self.assertEqual(manifest["selected_export_count"], 79)
            self.assertEqual(manifest["csv_count"], 79)
            self.assertIn("Individual_Fit_Data/Index.json", manifest["generated_files"])

    def test_duplicate_basename_different_folders_never_swaps_results(self):
        first, first_record = make_completed("run-1", "Run1_Output", source_path=r"D:\Run1\sample.asc", marker=1)
        second, second_record = make_completed("run-2", "Run2_Output", source_path=r"D:\Run2\sample.asc", marker=20)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, [first, second], {}, {}, records=[first_record, second_record], options=DIRECT_CSV)
            out = run_directory(report)
            first_csv = pd.read_csv(out / "Individual_Fit_Data" / "Run1_Output.csv")
            second_csv = pd.read_csv(out / "Individual_Fit_Data" / "Run2_Output.csv")
            np.testing.assert_allclose(first_csv["Fit_Input_Intensity"], first.result.y_raw)
            np.testing.assert_allclose(second_csv["Fit_Input_Intensity"], second.result.y_raw)
            self.assertFalse(np.allclose(first_csv["Fit_Input_Intensity"], second.result.y_raw))

    def test_duplicate_basename_first_failed_exports_only_second_result(self):
        failed_stub, failed_record = make_completed("run-1", "Failed_Run1", source_path=r"D:\Run1\sample.asc", marker=1)
        failed_record.update(status="Failed", failure_reason="optimizer failed")
        successful, successful_record = make_completed("run-2", "Successful_Run2", source_path=r"D:\Run2\sample.asc", marker=20)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, [successful], {}, {}, records=[failed_record, successful_record], options=DIRECT_CSV)
            out = run_directory(report)
            self.assertFalse((out / "Individual_Fit_Data" / "Failed_Run1.csv").exists())
            exported = pd.read_csv(out / "Individual_Fit_Data" / "Successful_Run2.csv")
            np.testing.assert_allclose(exported["Fit_Input_Intensity"], successful.result.y_raw)
            index = json.loads((out / "Individual_Fit_Data" / "Index.json").read_text(encoding="utf-8"))
            failed = next(row for row in index if row["Record_ID"] == "run-1")
            self.assertEqual(failed["Fit_Status"], "Failed")
            self.assertEqual(failed["CSV_Filename"], "")
            self.assertEqual(report["failed_fits"], 1)

    def test_duplicate_output_names_are_unique_but_record_ids_remain_distinct(self):
        first, first_record = make_completed("id-a", "sample")
        second, second_record = make_completed("id-b", "sample", marker=9)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, [first, second], {}, {}, records=[first_record, second_record], options=DIRECT_CSV)
            out = run_directory(report)
            self.assertEqual(sorted(path.name for path in (out / "Individual_Fit_Data").glob("*.csv")), ["sample.csv", "sample_2.csv"])
            index = json.loads((out / "Individual_Fit_Data" / "Index.json").read_text(encoding="utf-8"))
            self.assertEqual({row["Record_ID"] for row in index}, {"id-a", "id-b"})

    def test_output_name_changed_after_fit_uses_current_name_and_same_record_result(self):
        completed, record = make_completed("stable-id", "ABC", marker=17)
        record["code"] = "XYZ"
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, [completed], {}, {}, records=[record], options=DIRECT_CSV)
            out = run_directory(report)
            self.assertFalse((out / "Individual_Fit_Data" / "ABC.csv").exists())
            exported = pd.read_csv(out / "Individual_Fit_Data" / "XYZ.csv")
            np.testing.assert_allclose(exported["Fit_Input_Intensity"], completed.result.y_raw)

    def test_export_false_controls_every_spectrum_level_output(self):
        completed, records = [], []
        for record_id in ("A", "B", "C"):
            item, record = make_completed(record_id, record_id)
            completed.append(item)
            records.append(record)
        records[1]["export"] = False
        options = ExportOptions(figures_png=False, figures_pdf=False, timestamped_folder=False)
        temperature_fit = TemperatureFitResult(
            analysis_type="Peak energy",
            model="Linear",
            parameters={"slope": 1.0},
            errors={"slope": 0.1},
            ci95={"slope": (0.8, 1.2)},
            temperature_k=np.array([100.0, 200.0, 300.0]),
            observed=np.array([1.0, 2.0, 3.0]),
            fitted=np.array([1.0, 2.0, 3.0]),
            residual=np.zeros(3),
            r_squared=1.0,
            adjusted_r_squared=1.0,
            rmse=0.0,
            aic=0.0,
            aicc=0.0,
            bic=0.0,
            warnings=[],
        )
        tempfits = {"C1": {"Peak energy": temperature_fit}}
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, completed, tempfits, {}, records=records, options=options)
            out = run_directory(report)
            self.assertEqual(report["successful_curve_exports"], 2)
            self.assertEqual(report["skipped_by_export_checkbox"], 1)
            self.assertEqual(len(list((out / "Individual_Fit_Data").glob("*.csv"))), 2)
            workbook = load_workbook(out / "Individual_Fitting_Data.xlsx", read_only=True)
            self.assertEqual(len(workbook.sheetnames), 3)
            workbook.close()
            peaks = pd.read_excel(out / "PL_batch_results.xlsx", sheet_name="Peak_Results")
            metadata = pd.read_excel(out / "PL_batch_results.xlsx", sheet_name="Spectrum_Metadata")
            payload = json.loads((out / "analysis_results.json").read_text(encoding="utf-8"))
            self.assertEqual(set(peaks["Record_ID"]), {"A", "C"})
            self.assertEqual(set(metadata["Record_ID"]), {"A", "C"})
            self.assertEqual({row["record_id"] for row in payload["spectra"]}, {"A", "C"})
            self.assertIn("C1", payload["temperature_fits"])
            self.assertFalse(pd.read_csv(out / "temperature_fit_results.csv").empty)

    def test_79_successful_with_10_export_false_exports_exactly_69(self):
        completed, records = [], []
        for index in range(79):
            item, record = make_completed(f"select-{index}", f"Select_{index}")
            completed.append(item)
            records.append(record)
        for record in records[:10]:
            record["export"] = False
        options = ExportOptions(figures_png=False, figures_pdf=False, timestamped_folder=False)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, completed, {}, {}, records=records, options=options)
            out = run_directory(report)
            self.assertEqual(report["successful_curve_exports"], 69)
            self.assertEqual(report["csv_files"], 69)
            self.assertEqual(report["individual_excel_sheets"], 69)
            workbook = load_workbook(out / "Individual_Fitting_Data.xlsx", read_only=True)
            self.assertEqual(len(workbook.sheetnames), 70)
            workbook.close()
            peaks = pd.read_csv(out / "peak_results.csv")
            payload = json.loads((out / "analysis_results.json").read_text(encoding="utf-8"))
            self.assertEqual(peaks["Record_ID"].nunique(), 69)
            self.assertEqual(len(payload["spectra"]), 69)

    def test_individual_excel_and_index_contain_required_provenance(self):
        completed, record = make_completed("provenance-id", "Provenance", jacobian=True, n_peaks=2)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, [completed], {}, {}, records=[record], options=DIRECT_CURVES)
            out = run_directory(report)
            index = pd.read_excel(out / "Individual_Fitting_Data.xlsx", sheet_name="Index")
            sheet_name = index.loc[0, "Export_Sheet_Name"]
            raw_sheet = pd.read_excel(out / "Individual_Fitting_Data.xlsx", sheet_name=sheet_name, header=None)
            curve_header_index = raw_sheet.index[raw_sheet.iloc[:, 0] == "Photon_Energy_eV"][0]
            metadata = dict(zip(raw_sheet.iloc[:curve_header_index, 0], raw_sheet.iloc[:curve_header_index, 1]))
            required = {
                "Record_ID", "Output_Name", "Source_Filename", "Source_Full_Path",
                "Condition", "Sample_ID", "Temperature_K", "Model", "Number_of_Peaks",
                "Baseline_Mode", "Jacobian_Corrected", "Intensity_Domain", "Fit_Range_eV",
                "Instrument_FWHM_meV", "R2", "Adjusted_R2", "RMSE", "AICc", "BIC",
                "Peak_1_Energy_eV", "Peak_2_Energy_eV",
            }
            self.assertTrue(required.issubset(metadata))
            headers = raw_sheet.iloc[curve_header_index].dropna().tolist()
            self.assertIn("Fit_Input_Intensity", headers)
            self.assertIn("Peak_1_With_Baseline", headers)
            self.assertNotIn("Raw_Intensity", headers)

    def test_timestamped_runs_prevent_stale_files_and_preserve_user_file(self):
        rows_79, records_79 = [], []
        for index in range(79):
            item, record = make_completed(f"id-{index}", f"S{index}")
            rows_79.append(item)
            records_79.append(record)
        options = ExportOptions(summary_excel=False, individual_csv=True, individual_excel=False, figures_png=False, figures_pdf=False, timestamped_folder=True)
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            protected = base / "DO_NOT_DELETE.txt"
            protected.write_text("user data", encoding="utf-8")
            first = export_all(base, rows_79, {}, {}, records=records_79, options=options)
            second = export_all(base, rows_79[:77], {}, {}, records=records_79[:77], options=options)
            first_out, second_out = run_directory(first), run_directory(second)
            self.assertNotEqual(first_out, second_out)
            self.assertEqual(len(list((first_out / "Individual_Fit_Data").glob("*.csv"))), 79)
            self.assertEqual(len(list((second_out / "Individual_Fit_Data").glob("*.csv"))), 77)
            self.assertTrue(protected.is_file())

    def test_overwrite_mode_removes_only_manifest_files(self):
        first, first_record = make_completed("id-1", "first")
        second, second_record = make_completed("id-2", "second")
        with tempfile.TemporaryDirectory() as tmp:
            base = Path(tmp)
            protected = base / "DO_NOT_DELETE.txt"
            protected.write_text("user data", encoding="utf-8")
            export_all(base, [first, second], {}, {}, records=[first_record, second_record], options=DIRECT_CSV)
            export_all(base, [first], {}, {}, records=[first_record], options=DIRECT_CSV)
            self.assertEqual([path.name for path in (base / "Individual_Fit_Data").glob("*.csv")], ["first.csv"])
            self.assertTrue(protected.is_file())

    def test_jacobian_metadata_is_taken_from_saved_fit_settings(self):
        corrected, corrected_record = make_completed("jac-true", "corrected", jacobian=True)
        uncorrected, uncorrected_record = make_completed("jac-false", "uncorrected", jacobian=False)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, [corrected, uncorrected], {}, {}, records=[corrected_record, uncorrected_record], options=DIRECT_CSV)
            index = json.loads((run_directory(report) / "Individual_Fit_Data" / "Index.json").read_text(encoding="utf-8"))
            by_id = {row["Record_ID"]: row for row in index}
            self.assertTrue(by_id["jac-true"]["Jacobian_Corrected"])
            self.assertEqual(by_id["jac-true"]["Intensity_Domain"], "Energy-domain Jacobian-corrected")
            self.assertFalse(by_id["jac-false"]["Jacobian_Corrected"])
            self.assertEqual(by_id["jac-false"]["Intensity_Domain"], "Wavelength-domain input")

    def test_reserved_trailing_and_case_colliding_windows_names_are_safe(self):
        reserved = ["CON", "CON.txt", "AUX.dat", "NUL.foo", "COM1", "COM1.txt", "LPT9.dat"]
        for name in reserved:
            sanitized = sanitize_windows_filename(name)
            self.assertNotIn(sanitized.partition(".")[0].upper(), {"CON", "AUX", "NUL", "COM1", "LPT9"})
            self.assertFalse(sanitized.endswith((".", " ")))
        self.assertEqual(sanitize_windows_filename("sample. "), "sample")
        self.assertEqual(sanitize_windows_filename("sample..."), "sample")
        self.assertEqual(sanitize_windows_filename(" sample "), "sample")

        rows, records = [], []
        for index, name in enumerate(reserved + ["Sample", "sample", "SAMPLE"]):
            item, record = make_completed(f"safe-{index}", name)
            rows.append(item)
            records.append(record)
        with tempfile.TemporaryDirectory() as tmp:
            report = export_all(tmp, rows, {}, {}, records=records, options=DIRECT_CSV)
            names = [path.name for path in (run_directory(report) / "Individual_Fit_Data").glob("*.csv")]
            self.assertEqual(len(names), len({name.casefold() for name in names}))
            self.assertEqual(len(names), len(rows))

    def test_summary_workbook_omits_curve_sheets_unless_legacy_enabled(self):
        completed, record = make_completed("id-1", "sample")
        with tempfile.TemporaryDirectory() as tmp:
            normal = ExportOptions(individual_csv=False, individual_excel=False, figures_png=False, figures_pdf=False, timestamped_folder=False)
            export_all(tmp, [completed], {}, {}, records=[record], options=normal)
            workbook = load_workbook(Path(tmp) / "PL_batch_results.xlsx", read_only=True)
            self.assertEqual(workbook.sheetnames, ["Peak_Results", "Temperature_Fits", "Spectrum_Metadata"])
            workbook.close()
            legacy = ExportOptions(individual_csv=False, individual_excel=False, figures_png=False, figures_pdf=False, legacy_curve_sheets=True, timestamped_folder=False)
            export_all(tmp, [completed], {}, {}, records=[record], options=legacy)
            workbook = load_workbook(Path(tmp) / "PL_batch_results.xlsx", read_only=True)
            self.assertTrue(any(name.startswith("Fit_") for name in workbook.sheetnames))
            self.assertTrue(any(name.startswith("Res_") for name in workbook.sheetnames))
            workbook.close()


if __name__ == "__main__":
    unittest.main()
